from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _dt(value):
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M")


def build_tasks_export_excel(tasks: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = [
        "ID",
        "Название",
        "Статус",
        "Приоритет",
        "Постановщик",
        "Исполнитель",
        "Создана",
        "Дедлайн",
        "Описание",
        "Последний комментарий",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for task in tasks:
        creator = "—"
        if getattr(task, "creator", None):
            creator = task.creator.full_name or (f"@{task.creator.username}" if task.creator.username else str(task.creator.tg_id))

        assignee = "—"
        if getattr(task, "assignee", None):
            assignee = task.assignee.full_name or (f"@{task.assignee.username}" if task.assignee.username else str(task.assignee.tg_id))

        status = task.status.label if task.status else ""
        priority = task.priority.label if task.priority else ""

        last_comment = ""

        if getattr(task, "comments", None):
            last = max(task.comments, key=lambda c: c.created_at, default=None)
            if last:
                last_comment = last.text
        last_comment = (last_comment or "")[:1000]


        ws.append([
            task.id,
            task.title or "",
            status,
            priority,
            creator,
            assignee,
            _dt(task.created_at),
            _dt(task.deadline_at),
            task.description or "",
            last_comment,
        ])

    widths = {
        1: 8,
        2: 30,
        3: 18,
        4: 16,
        5: 24,
        6: 24,
        7: 20,
        8: 20,
        9: 50,
        10: 50,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


